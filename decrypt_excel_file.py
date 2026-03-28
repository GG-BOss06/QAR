import requests
import json

# 读取加密结果文件
with open('encrypted_qar_excel_result.json', 'r', encoding='utf-8') as f:
    encryption_result = json.load(f)

# 准备解密请求数据
payload = {
    'encryptedData': encryption_result['encryptedData'],
    'policy': 'role:admin AND department:security'
}

print("正在解密Excel文件数据...")
print(f"原始文件名: {encryption_result['originalFileName']}")
print(f"文件大小: {encryption_result['fileSize']} bytes")
print(f"加密数据长度: {len(encryption_result['encryptedData'])} 字符")
print(f"发送解密请求到: http://localhost:8080/api/decrypt-excel")

try:
    # 发送POST请求
    response = requests.post('http://localhost:8080/api/decrypt-excel', json=payload, timeout=30)
    print(f"响应状态码: {response.status_code}")
    print(f"响应内容: {response.text[:500]}...")
    
    response.raise_for_status()
    
    result = response.json()
    
    print("\n" + "="*50)
    print("解密结果")
    print("="*50)
    print(f"状态码: {result['code']}")
    print(f"消息: {result['message']}")
    print(f"\n解密后数据预览 (前500字符):")
    print(result['preview'])
    print(f"\n解密后数据长度: {len(result['decryptedData'])} 字符")
    print("="*50)
    
    # 保存解密结果到文件
    with open('decrypted_qar_excel_result.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print("\n解密结果已保存到: decrypted_qar_excel_result.json")
    
    # 保存解密后的数据为Excel文件
    print("\n正在将解密后的数据保存为Excel文件...")
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # 解析解密后的数据
    lines = result['decryptedData'].strip().split('\n')
    sheet_name = None
    
    for line in lines:
        line = line.strip()
        if line.startswith('=== Sheet:'):
            # 新的工作表
            sheet_name = line.split('=== Sheet: ')[1].split(' ===')[0]
            if sheet_name != ws.title:
                # 创建新的工作表
                ws = wb.create_sheet(title=sheet_name)
        elif line and not line.startswith('==='):
            # 数据行
            cells = line.split('\t')
            ws.append(cells)
    
    # 删除默认的Sheet工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 保存Excel文件
    wb.save('newexcel.xlsx')
    print("Excel文件已保存为: newexcel.xlsx")
    
except Exception as e:
    print(f"错误: {e}")
    import traceback
    traceback.print_exc()
    if hasattr(e, 'response') and e.response:
        print(f"响应状态码: {e.response.status_code}")
        print(f"响应内容: {e.response.text}")
