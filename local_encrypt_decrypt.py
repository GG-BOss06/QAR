import base64
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import os
import openpyxl

# 固定的测试密钥
KEY = b'12345678901234567890123456789012'

# 加密函数
def encrypt_data(data):
    # 生成随机IV
    iv = os.urandom(12)
    # 创建加密器
    encryptor = Cipher(
        algorithms.AES(KEY),
        modes.GCM(iv),
        backend=default_backend()
    ).encryptor()
    # 加密数据
    ciphertext = encryptor.update(data.encode('utf-8')) + encryptor.finalize()
    # 返回IV和密文的Base64编码
    return base64.b64encode(iv + ciphertext + encryptor.tag).decode('utf-8')

# 解密函数
def decrypt_data(encrypted_data):
    # 解码Base64数据
    data = base64.b64decode(encrypted_data)
    # 提取IV、密文和tag
    iv = data[:12]
    tag = data[-16:]
    ciphertext = data[12:-16]
    # 创建解密器
    decryptor = Cipher(
        algorithms.AES(KEY),
        modes.GCM(iv, tag),
        backend=default_backend()
    ).decryptor()
    # 解密数据
    plaintext = decryptor.update(ciphertext) + decryptor.finalize()
    return plaintext.decode('utf-8')

# 读取Excel文件内容
def read_excel_content(file_path):
    wb = openpyxl.load_workbook(file_path)
    content = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        content.append(f"=== Sheet: {sheet_name} ===")
        for row in sheet.iter_rows(values_only=True):
            row_content = '\t'.join(str(cell) if cell is not None else '' for cell in row)
            if row_content.strip():
                content.append(row_content)
        content.append('')
    return '\n'.join(content)

# 将内容保存为Excel文件
def save_content_to_excel(content, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    lines = content.strip().split('\n')
    sheet_name = None
    
    for line in lines:
        line = line.strip()
        if line.startswith('=== Sheet:'):
            # 新的工作表
            sheet_name = line.split('=== Sheet: ')[1].split(' ===')[0]
            if sheet_name != ws.title:
                # 创建新的工作表
                ws = wb.create_sheet(title=sheet_name)
        elif line and not line.startswith('==='):
            # 数据行
            cells = line.split('\t')
            ws.append(cells)
    
    # 删除默认的Sheet工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"Excel文件已保存为: {output_file}")

# 主函数
def main():
    print("=== Excel文件加密和解密工具 ===")
    
    # 加密Excel文件
    print("\n1. 加密Excel文件")
    input_file = "QAR示例数据.xlsx"
    print(f"输入文件: {input_file}")
    
    # 读取Excel内容
    excel_content = read_excel_content(input_file)
    print(f"Excel内容读取完成，长度: {len(excel_content)} 字符")
    
    # 加密数据
    encrypted_data = encrypt_data(excel_content)
    print(f"加密完成，加密数据长度: {len(encrypted_data)} 字符")
    
    # 保存加密结果
    with open('encrypted_result.json', 'w', encoding='utf-8') as f:
        f.write(encrypted_data)
    print("加密结果已保存到: encrypted_result.json")
    
    # 解密数据
    print("\n2. 解密Excel文件")
    with open('encrypted_result.json', 'r', encoding='utf-8') as f:
        encrypted_data = f.read()
    
    decrypted_content = decrypt_data(encrypted_data)
    print(f"解密完成，解密数据长度: {len(decrypted_content)} 字符")
    
    # 保存解密结果为Excel文件
    output_file = "newexcel.xlsx"
    save_content_to_excel(decrypted_content, output_file)
    
    print("\n=== 操作完成 ===")

if __name__ == "__main__":
    main()
